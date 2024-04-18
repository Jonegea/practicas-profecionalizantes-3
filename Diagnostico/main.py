from openpyxl import load_workbook

def read_data_from_excel(file_path):
    set_datos = set()

    wb = load_workbook(filename=file_path)
    sheet = wb.active
    #utilizo row para iterar las filas
    for row in sheet.iter_rows(min_row=4, values_only=True):
        nombre = row[0]
        municipio = row[5]
        departamento = row[6]
        provincia = row[7]
        set_datos.add((nombre, municipio, departamento, provincia))
    
    return set_datos

# cargo o leo el archivo excel
if __name__ == "__main__":
    file_path = "Localidades.xlsx"
    datos = read_data_from_excel(file_path)
    for nombre, municipio , departamento, provincia in datos:
        print(f"Ciudad : {nombre}, Municipio : {municipio}, Departamento: {departamento}, Provincia: {provincia}")
